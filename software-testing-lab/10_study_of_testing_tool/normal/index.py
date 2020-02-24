import openpyxl
wb=openpyxl.load_workbook("Avengers.xlsx")
print(wb.sheetnames)
print(wb.active.title)
sh=wb["Sheet1"]
print(sh['A3'].value)
c1=sh.cell(1,1)
print(c1.value)
rows=sh.max_row
print(rows)
cols=sh.max_column
print(cols)
wb=openpyxl.Workbook()
print(wb.sheetnames)
sh=wb.active
sh['A1'].value="VK"
wb.create_sheet("Hello")
wb.save("Revengers.xlsx")


